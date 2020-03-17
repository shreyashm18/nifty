import datetime, calendar
import openpyxl
from database_connection import db_connection_execute
import mysql.connector
from script_for_calculating_MSF import msf_f

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="root",
  database="martha",
  #auth_plugin="mysql_native_password"
)

cursor=mydb.cursor()

wb=openpyxl.load_workbook('demo.xlsx')
sht1=wb.get_sheet_by_name('Sheet1')
max_rows=sht1.max_row
max_columns=sht1.max_column

# data=sht1.cell(row=6,column=1).value
# print(data)
# day = data.weekday()
# print(day)
day_dict={0:'Mon',1:'Tue',2:'Wed',3:'Thu',4:'Fri',5:'Sat',6:'Sun'}

# day=day_dict[day]
# print(day)
sr=1



for row in range(1,max_rows+1):
	Nifty=Sr=Open=High=Low=Close=''
	
	if row<5:
			continue
	else:
		for column in range(1,6+1):
		
			data=sht1.cell(row=row,column=column).value
			# print(data)

			
			if not Nifty:
				Nifty=data
				day=Nifty.weekday()
				day=day_dict[day]

			elif not Sr:
				Sr=data
			
			elif not Open:
				Open=data
			elif not High:
				High=data
			elif not Low:
				Low=data
			else:
				Close=data

		Sr=sr
		sr+=1
		query = "INSERT INTO martha.nifty50 (`Date`,`day`,Sr,Open,High,Low,Close) VALUES (%s,%s,%s,%s,%s,%s,%s)"
		values=(Nifty,day,Sr,Open,High,Low,Close)
		#cursor.execute(sql)
		db_connection_execute("localhost", "root", "root", "martha", query, values = values)
		mydb.commit()


msf_f()
