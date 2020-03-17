import mysql.connector
import openpyxl

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="root",
  database="martha",
  #auth_plugin="mysql_native_password"
)
mycursor = mydb.cursor()

wb=openpyxl.load_workbook('demo.xlsx')
sht1=wb.get_sheet_by_name('Sheet1')
max_rows=sht1.max_row
max_columns=sht1.max_column
#Nifty=Sr=Open=''
#for i in range(1,7):
# li=[]
# li.append(str(sht1.cell(row=1,column=1).value))
# li.append(sht1.cell(row=1,column=2).value)
# li.append(str(sht1.cell(row=1,column=3).value))
# li.append(str(sht1.cell(row=1,column=4).value))
# li.append(str(sht1.cell(row=1,column=4).value))
# li.append(str(sht1.cell(row=1,column=5).value))
# print(Nifty)
# print(Sr)
# print(Open)
#li=[Nifty,Sr,Open,High,Low,Close]
#sql="CREATE TABLE nifty50(Nifty varchar(255),Sr int,Open varchar(255),High varchar(255),Low varchar(255),Close varchar(255))"
#mycursor.execute(sql)	
#mycursor.commit()

for i in range(5,max_rows+1):
	
	Nifty=sht1.cell(row=i,column=1).value
	Sr=sht1.cell(row=i,column=2).value
	Open=sht1.cell(row=i,column=3).value
	High=str(sht1.cell(row=i,column=4).value)
	Low=str(sht1.cell(row=i,column=5).value)
	Close=str(sht1.cell(row=i,column=6).value)
	# if i==1:
	# 	sql='CREATE TABLE nifty50 (f"{Nifty} varchar(255),{Sr} int,{Open} VARCHAR(255),{High} VARCHAR(255),{Low} VARCHAR(255),{Close} VARCHAR(255)")'
	# 	mycursor.execute(sql)
	# 	mycursor.commit()
	# 	#print(f'{Nifty} nifty50')		
	
	
	sql="INSERT INTO nifty50 (Nifty, Sr, Open, High, Low, Close) VALUES(%s,%s,%s,%s,%s,%s)"
	values=Nifty,Sr,Open,High,Low,Close
	mycursor.execute(sql,values)
	mydb.commit()
	#print(Sr)
	#print(Open)
