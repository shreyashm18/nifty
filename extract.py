import mysql.connector
import openpyxl

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="root",
  database="martha",
  #auth_plugin="mysql_native_password"
)
mycursor = mydb.cursor()

wb=openpyxl.load_workbook('emp.xlsx')
sht1=wb.get_sheet_by_name('Sheet1')
max_rows=sht1.max_row
max_columns=sht1.max_column
for i in range(2,max_rows+1):
	# empid=''
	# name=''
	# surname=''
	# age=''
	#for j in range(1,max_columns+1):
		# cell_obj=sht1.cell(row=i,column=j).value
		# if not empid:
		# 	empid=cell_obj
		# elif not name:
		# 	name=str(cell_obj)
		# elif not surname:
		# 	surname=str(cell_obj)
		# elif not age:
		# 	age=cell_obj
		#print(cell_obj)
	empid=sht1.cell(row=i,column=1).value
	name=sht1.cell(row=i,column=2).value
	surname=sht1.cell(row=i,column=3).value
	age=sht1.cell(row=i,column=4).value

	sql = "INSERT INTO employee (id, name, surname, age) VALUES (%s,%s,%s,%s)"
	values=empid,name,surname,age
	mycursor.execute(sql,values)
	mydb.commit()
print(mycursor.rowcount," record inserted")

